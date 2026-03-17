
import openpyxl
import json
from datetime import datetime

def extract_budget_data(file_path):
    wb = openpyxl.load_workbook(file_path, data_only=True)
    sheet = wb['BALANCE']
    
    data = {
        "2025-01": {"income": [], "expenses": []},
        "2025-02": {"income": [], "expenses": []}
    }
    
    # Income mapping (Rows 3-6)
    for row in range(3, 7):
        # JAN
        name = sheet.cell(row=row, column=1).value
        est = sheet.cell(row=row, column=3).value or 0
        act = sheet.cell(row=row, column=4).value or 0
        if name:
            data["2025-01"]["income"].append({"name": name, "estimated": float(est), "actual": float(act)})
        # FEB
        name_feb = sheet.cell(row=row, column=10).value
        est_feb = sheet.cell(row=row, column=12).value or 0
        act_feb = sheet.cell(row=row, column=13).value or 0
        if name_feb:
            data["2025-02"]["income"].append({"name": name_feb, "estimated": float(est_feb), "actual": float(act_feb)})

    # Expenses mapping (Looking deeper into rows)
    # January expenses are in cols 1-4, starting from around row 23 down to 40
    # February expenses are in cols 10-13, starting from row 23
    for row in range(23, 50):
        # JAN
        name = sheet.cell(row=row, column=1).value
        date_raw = sheet.cell(row=row, column=2).value
        act = sheet.cell(row=row, column=3).value or 0
        est = sheet.cell(row=row, column=4).value or 0
        if name and name != "None" and isinstance(name, str):
            date_str = date_raw.strftime("%Y-%m-%d") if isinstance(date_raw, datetime) else str(date_raw)
            data["2025-01"]["expenses"].append({
                "name": name, 
                "date": date_str, 
                "estimated": abs(float(est)), 
                "actual": abs(float(act))
            })
            
        # FEB
        name_feb = sheet.cell(row=row, column=10).value
        date_raw_feb = sheet.cell(row=row, column=11).value
        act_feb = sheet.cell(row=row, column=12).value or 0
        est_feb = sheet.cell(row=row, column=13).value or 0
        if name_feb and name_feb != "None" and isinstance(name_feb, str):
            date_str_feb = date_raw_feb.strftime("%Y-%m-%d") if isinstance(date_raw_feb, datetime) else str(date_raw_feb)
            data["2025-02"]["expenses"].append({
                "name": name_feb, 
                "date": date_str_feb, 
                "estimated": abs(float(est_feb)), 
                "actual": abs(float(act_feb))
            })
            
    return data

if __name__ == "__main__":
    file_path = r"C:\Users\pedro\OneDrive\Github\Budget\Budget 2026.xlsx"
    try:
        data = extract_budget_data(file_path)
        with open('import_data.json', 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=2)
        print("Successfully updated import_data.json")
    except Exception as e:
        print(f"Error: {e}")
